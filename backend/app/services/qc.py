import io
import json
import math
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from scipy import stats
from sklearn.decomposition import PCA as PCA_SKL
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app import models, schemas
from app.services.preprocessing import to_dataframe, _to_json_safe
from app.services.plots import _merge_style, _group_color_map, _apply_base_layout, generate_plot, _shorten_name


def _safe_log10(values):
    with np.errstate(divide="ignore", invalid="ignore"):
        return np.log10(np.where(np.array(values) > 0, np.array(values), np.nan))


def _as_floats(df: pd.DataFrame) -> pd.DataFrame:
    return df.apply(pd.to_numeric, errors="coerce")


def _compute_mahalanobis_outliers(scores: np.ndarray, quantile: float = 0.99) -> np.ndarray:
    if len(scores) < 3:
        return np.zeros(len(scores), dtype=bool)
    mean = scores.mean(axis=0)
    cov = np.cov(scores, rowvar=False)
    if cov.ndim < 2 or cov.shape[0] < 2:
        cov = np.atleast_2d(np.var(scores, axis=0, ddof=1))
    try:
        inv_cov = np.linalg.inv(cov)
    except np.linalg.LinAlgError:
        return np.zeros(len(scores), dtype=bool)
    centered = scores - mean
    distances = np.array([x @ inv_cov @ x for x in centered])
    # chi2 threshold for 2 df at given quantile
    threshold = stats.chi2.ppf(quantile, df=2)
    return distances > threshold


def _filter_by_groups(df: pd.DataFrame, sample_meta: dict, selected_groups: list | None) -> tuple[pd.DataFrame, dict]:
    if not selected_groups:
        return df, sample_meta
    # Normalize whitespace on both sides so a leading/trailing space in the stored metadata
    # does not cause the selection to silently fall back to the full dataset.
    selected_set = {str(g).strip() for g in selected_groups}
    norm_meta = {c: str(sample_meta.get(c, "Unknown")).strip() for c in df.columns}
    selected_cols = [c for c in df.columns if norm_meta[c] in selected_set]
    if not selected_cols:
        return df.iloc[:, 0:0], {}
    return df[selected_cols], {c: str(sample_meta[c]).strip() for c in selected_cols}


def _build_filtered_dataset(dataset: models.Dataset, df: pd.DataFrame, sample_meta: dict) -> models.Dataset:
    return models.Dataset(
        id=dataset.id,
        project_id=dataset.project_id,
        source_file_id=dataset.source_file_id,
        name=dataset.name,
        feature_type=dataset.feature_type,
        data_matrix=df.to_dict("list"),
        sample_metadata=sample_meta,
        feature_metadata=list(dataset.feature_metadata or []),
        processing_history=list(dataset.processing_history or []),
    )


def qc_analysis(dataset: models.Dataset, style: dict | None = None, selected_groups: list | None = None) -> dict:
    style = _merge_style(style)
    df = _as_floats(to_dataframe(dataset))
    sample_meta = dataset.sample_metadata or {}
    feature_meta = dataset.feature_metadata or []
    df, sample_meta = _filter_by_groups(df, sample_meta, selected_groups)

    samples = list(df.columns)
    groups = [sample_meta.get(s, "Unknown") for s in samples]
    group_order = sorted(set(groups))
    if "Unknown" in group_order:
        group_order = [g for g in group_order if g != "Unknown"] + ["Unknown"]
    color_map = _group_color_map(style, group_order)

    # Basic metrics
    num_features, num_samples = df.shape
    missing = df.isna().sum()
    missing_pct = (missing / num_features * 100).round(2).to_dict()
    total_missing_pct = round(float(df.isna().sum().sum() / df.size * 100), 2)

    tic = df.sum(numeric_only=True, skipna=True).round(2).to_dict()
    log2_tic = {s: round(float(np.log2(v)) if v and v > 0 else 0.0, 2) for s, v in tic.items()}
    detected = (df.notna().sum(axis=0)).to_dict()

    # Group CV (median per-feature CV within each group)
    group_cvs = {}
    for g in group_order:
        cols = [s for s in samples if sample_meta.get(s, "Unknown") == g]
        if not cols:
            continue
        sub = df[cols]
        with np.errstate(divide="ignore", invalid="ignore"):
            mean = sub.mean(axis=1, skipna=True)
            std = sub.std(axis=1, skipna=True)
            cv = (std / mean).replace([np.inf, -np.inf], np.nan).dropna()
        group_cvs[g] = round(float(cv.median()) * 100, 2) if not cv.empty else None

    qc_groups = {g for g in group_order if "qc" in g.lower() or "quality" in g.lower()}
    blank_groups = {g for g in group_order if any(b in g.lower() for b in ["blank", "solvent", "ntc", "standard", "pool"])}

    qc_median_cv = None
    if qc_groups:
        qc_cols = [s for s in samples if sample_meta.get(s, "Unknown") in qc_groups]
        if qc_cols:
            sub = df[qc_cols]
            with np.errstate(divide="ignore", invalid="ignore"):
                mean = sub.mean(axis=1, skipna=True)
                std = sub.std(axis=1, skipna=True)
                cv = (std / mean).replace([np.inf, -np.inf], np.nan).dropna()
            qc_median_cv = round(float(cv.median()) * 100, 2) if not cv.empty else None

    blank_sample_ratio = None
    if blank_groups:
        blank_cols = [s for s in samples if sample_meta.get(s, "Unknown") in blank_groups]
        sample_cols = [s for s in samples if sample_meta.get(s, "Unknown") not in blank_groups]
        if blank_cols and sample_cols:
            blank_mean = df[blank_cols].mean(axis=1, skipna=True).replace(0, np.nan)
            sample_mean = df[sample_cols].mean(axis=1, skipna=True)
            ratio = (sample_mean / blank_mean).replace([np.inf, -np.inf], np.nan).dropna()
            blank_sample_ratio = round(float(ratio.median()), 2) if not ratio.empty else None

    # PCA outlier detection (log10 + scale)
    pca_df = df.copy().replace(0, np.nan)
    log_df = np.log10(pca_df)
    log_df = log_df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    if log_df.shape[0] >= 2 and log_df.shape[1] >= 2:
        log_df = log_df.fillna(log_df.min().min() / 2)
        X = StandardScaler().fit_transform(log_df.T)
        n_comp = min(2, X.shape[0], X.shape[1])
        pca = PCA_SKL(n_components=n_comp)
        scores = pca.fit_transform(X)
        if scores.shape[1] >= 2:
            outliers = _compute_mahalanobis_outliers(scores)
            outlier_count = int(outliers.sum())
            outlier_samples = [samples[i] for i, flag in enumerate(outliers) if flag]
        else:
            outlier_count = 0
            outlier_samples = []
    else:
        outlier_count = 0
        outlier_samples = []

    # Helper for Plotly bar colored by group
    def _bar_figure(title: str, y: dict, y_title: str) -> go.Figure:
        fig = go.Figure()
        for g in group_order:
            xs = [s for s in samples if sample_meta.get(s, "Unknown") == g]
            ys = [y.get(s, 0) for s in xs]
            fig.add_trace(go.Bar(x=xs, y=ys, name=g, marker_color=color_map.get(g, "#94a3b8")))
        fig.update_layout(barmode="group", xaxis_title="Sample", yaxis_title=y_title)
        _apply_base_layout(fig, style, title=title, x_labels=samples)
        fig.update_xaxes(tickangle=45)
        return json.loads(fig.to_json())

    # Box plot of log2 intensities per sample
    def _log2_box_figure() -> go.Figure:
        fig = go.Figure()
        for g in group_order:
            xs = [s for s in samples if sample_meta.get(s, "Unknown") == g]
            for s in xs:
                vals = _safe_log10(df[s].dropna().values)
                vals = vals[~np.isnan(vals)]
                if len(vals) == 0:
                    continue
                fig.add_trace(go.Box(y=vals, name=s, marker_color=color_map.get(g, "#94a3b8"), showlegend=False))
        fig.update_layout(xaxis_title="Sample", yaxis_title="log2 intensity")
        _apply_base_layout(fig, style, title="Sample Intensity Distribution", x_labels=samples)
        fig.update_xaxes(tickangle=45)
        return json.loads(fig.to_json())

    # CV box plot per group
    def _cv_box_figure() -> go.Figure:
        fig = go.Figure()
        for g in group_order:
            cols = [s for s in samples if sample_meta.get(s, "Unknown") == g]
            if not cols:
                continue
            sub = df[cols]
            with np.errstate(divide="ignore", invalid="ignore"):
                mean = sub.mean(axis=1, skipna=True)
                std = sub.std(axis=1, skipna=True)
                cvs = (std / mean * 100).replace([np.inf, -np.inf], np.nan).dropna().values
            cvs = cvs[cvs > 0]
            if len(cvs) == 0:
                continue
            fig.add_trace(go.Box(y=cvs, name=g, marker_color=color_map.get(g, "#94a3b8")))
        fig.update_layout(xaxis_title="Group", yaxis_title="Coefficient of variation (%)")
        _apply_base_layout(fig, style, title="Per-Feature CV by Group")
        return json.loads(fig.to_json())

    figures = {
        "tic": _bar_figure("Total Ion Current (TIC) by Sample", tic, "TIC"),
        "missing_pct": _bar_figure("Missing Values per Sample (%)", missing_pct, "Missing %"),
        "detected_features": _bar_figure("Detected Features per Sample", detected, "Count"),
        "log2_intensity": _log2_box_figure(),
        "cv_by_group": _cv_box_figure(),
    }

    filtered_dataset = _build_filtered_dataset(dataset, df, sample_meta)

    try:
        figures["pca"] = generate_plot(
            filtered_dataset,
            schemas.PlotRequest(plot_type="pca", parameters={"plot": "score"}, style=style),
        )
    except Exception:
        pass

    try:
        figures["correlation_heatmap"] = generate_plot(
            filtered_dataset,
            schemas.PlotRequest(plot_type="heatmap", parameters={"heatmap_type": "correlation"}, style=style),
        )
    except Exception:
        pass

    return {
        "metrics": {
            "num_features": num_features,
            "num_samples": num_samples,
            "num_groups": len(group_order),
            "group_counts": {g: groups.count(g) for g in group_order},
            "total_missing_pct": total_missing_pct,
            "missing_per_sample": missing_pct,
            "tic": tic,
            "log2_tic": log2_tic,
            "detected_features": detected,
            "group_cv_pct": group_cvs,
            "qc_median_cv_pct": qc_median_cv,
            "sample_to_blank_median_ratio": blank_sample_ratio,
            "pca_outlier_count": outlier_count,
            "pca_outlier_samples": outlier_samples,
        },
        "figures": figures,
    }


def qc_export_excel(dataset: models.Dataset, style: dict | None = None, selected_groups: list | None = None) -> bytes:
    """Build a styled, color-coded Excel QC summary workbook."""
    result = qc_analysis(dataset, style, selected_groups)
    metrics = result["metrics"]
    sample_meta = dataset.sample_metadata or {}
    full_df = _as_floats(to_dataframe(dataset))
    _, sample_meta = _filter_by_groups(full_df, sample_meta, selected_groups)
    style = _merge_style(style)

    wb = Workbook()
    # Remove the default sheet and recreate it with a proper name later.
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1e293b")
    good_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")  # green
    warn_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")  # amber
    bad_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")  # red
    thin_side = Side(style="thin", color="e2e8f0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _write_table(ws, title, headers, rows, col_widths=None):
        ws.append([title])
        ws.cell(row=1, column=1).font = title_font
        ws.append([])
        ws.append(headers)
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for row in rows:
            ws.append(row)
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = w

    # Overview sheet
    ws = wb.create_sheet("Overview")
    qc_cv = metrics.get("qc_median_cv_pct")
    blank_ratio = metrics.get("sample_to_blank_median_ratio")
    total_missing = metrics.get("total_missing_pct", 0)
    outlier_count = metrics.get("pca_outlier_count", 0)

    overview_rows = [
        ["Total features", metrics["num_features"], ""],
        ["Total samples", metrics["num_samples"], ""],
        ["Groups", metrics["num_groups"], ""],
        ["Total missing %", f"{total_missing}%", _status(total_missing, 20, 40)],
        ["QC median CV %", f"{qc_cv}%" if qc_cv is not None else "N/A", _status(qc_cv if qc_cv is not None else 0, 20, 30, invert=False) if qc_cv is not None else ""],
        ["Sample/Blank median ratio", blank_ratio if blank_ratio is not None else "N/A", _status(blank_ratio if blank_ratio is not None else 0, 10, 5, invert=True) if blank_ratio is not None else ""],
        ["PCA outlier count", outlier_count, _status(outlier_count, 0, 1, invert=False)],
    ]
    for g, cnt in metrics.get("group_counts", {}).items():
        cv = metrics.get("group_cv_pct", {}).get(g)
        overview_rows.append([f"Group {g} count", cnt, ""])
        if cv is not None:
            overview_rows.append([f"Group {g} median CV %", f"{cv}%", _status(cv, 20, 30)])

    _write_table(ws, "QC Summary", ["Metric", "Value", "Status"], overview_rows, [40, 20, 15])
    for row in range(4, ws.max_row + 1):
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center")
        status = ws.cell(row=row, column=3).value
        if status == "PASS":
            ws.cell(row=row, column=3).fill = good_fill
        elif status == "WARN":
            ws.cell(row=row, column=3).fill = warn_fill
        elif status == "FAIL":
            ws.cell(row=row, column=3).fill = bad_fill
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

    # Per-sample sheet
    ws2 = wb.create_sheet("Per Sample")
    samples = [s for s in metrics["tic"].keys()]
    outlier_set = set(metrics.get("pca_outlier_samples", []))
    per_sample_rows = []
    for s in samples:
        g = sample_meta.get(s, "Unknown")
        per_sample_rows.append([
            _shorten_name(s),
            g,
            metrics["tic"].get(s),
            metrics["log2_tic"].get(s),
            f"{metrics['missing_per_sample'].get(s, 0)}%",
            metrics["detected_features"].get(s, 0),
            "Yes" if s in outlier_set else "No",
        ])
    _write_table(ws2, "Per-Sample QC Metrics", ["Sample", "Group", "TIC", "log2 TIC", "Missing %", "Detected Features", "Outlier"], per_sample_rows, [24, 16, 16, 16, 14, 18, 12])
    gcolor_map = _group_color_map(style, sorted(set(sample_meta.values())))
    for row in range(4, ws2.max_row + 1):
        for col in range(1, 8):
            ws2.cell(row=row, column=col).border = thin_border
        # Color group cell by group color
        g_val = ws2.cell(row=row, column=2).value
        if g_val and g_val in gcolor_map:
            hex_color = gcolor_map[g_val].lstrip("#")
            ws2.cell(row=row, column=2).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            ws2.cell(row=row, column=2).font = Font(color="FFFFFF" if _is_dark(hex_color) else "000000")
        # Missing % color coding
        try:
            missing_val = float(ws2.cell(row=row, column=5).value.rstrip("%"))
            if missing_val > 50:
                ws2.cell(row=row, column=5).fill = bad_fill
            elif missing_val > 20:
                ws2.cell(row=row, column=5).fill = warn_fill
            else:
                ws2.cell(row=row, column=5).fill = good_fill
        except Exception:
            pass
        if ws2.cell(row=row, column=7).value == "Yes":
            ws2.cell(row=row, column=7).fill = bad_fill

    # Group CV sheet
    ws3 = wb.create_sheet("Group CV")
    cv_rows = []
    group_counts = metrics.get("group_counts", {})
    for g in sorted(group_counts):
        cv = metrics.get("group_cv_pct", {}).get(g)
        cv_rows.append([g, group_counts[g], f"{cv}%" if cv is not None else "N/A", _status(cv, 20, 30) if cv is not None else ""])
    _write_table(ws3, "Per-Group Coefficient of Variation", ["Group", "Sample Count", "Median CV %", "Status"], cv_rows, [20, 16, 16, 12])
    for row in range(4, ws3.max_row + 1):
        for col in range(1, 5):
            ws3.cell(row=row, column=col).border = thin_border
        status = ws3.cell(row=row, column=4).value
        if status == "PASS":
            ws3.cell(row=row, column=4).fill = good_fill
        elif status == "WARN":
            ws3.cell(row=row, column=4).fill = warn_fill
        elif status == "FAIL":
            ws3.cell(row=row, column=4).fill = bad_fill

    # Outliers sheet
    ws4 = wb.create_sheet("PCA Outliers")
    outlier_rows = []
    for s in metrics.get("pca_outlier_samples", []):
        outlier_rows.append([_shorten_name(s), sample_meta.get(s, "Unknown")])
    _write_table(ws4, "PCA Mahalanobis Outliers", ["Sample", "Group"], outlier_rows, [24, 16])
    for row in range(4, ws4.max_row + 1):
        for col in range(1, 3):
            ws4.cell(row=row, column=col).border = thin_border
            ws4.cell(row=row, column=col).fill = bad_fill

    # Notes sheet
    ws5 = wb.create_sheet("Thresholds")
    notes = [
        ["Metric", "PASS", "WARN", "FAIL"],
        ["Missing % per sample", "<= 20%", "20-50%", "> 50%"],
        ["QC median CV %", "<= 20%", "20-30%", "> 30%"],
        ["Group median CV %", "<= 20%", "20-30%", "> 30%"],
        ["Sample/Blank median ratio", ">= 10", "5-10", "< 5"],
        ["PCA outlier count", "0", "1", ">= 2"],
    ]
    _write_table(ws5, "QC Thresholds", notes[0], notes[1:], [32, 14, 14, 14])
    for row in range(4, ws5.max_row + 1):
        for col in range(1, 5):
            ws5.cell(row=row, column=col).border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _status(value, warn, fail, invert: bool = False):
    """Return PASS/WARN/FAIL based on numeric thresholds."""
    if value is None:
        return ""
    try:
        value = float(value)
    except (TypeError, ValueError):
        return ""
    if invert:
        if value < fail:
            return "FAIL"
        if value < warn:
            return "WARN"
        return "PASS"
    if value > fail:
        return "FAIL"
    if value > warn:
        return "WARN"
    return "PASS"


def _is_dark(hex_color: str) -> bool:
    """Return True for dark hex colors so white font can be used."""
    try:
        h = hex_color.lstrip("#")
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        return luminance < 0.5
    except Exception:
        return False

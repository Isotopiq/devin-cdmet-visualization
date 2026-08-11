import os
import re
from typing import Any, Dict, List, Tuple
import math

import openpyxl
import pandas as pd
from app.config import settings
from app.services import storage
import logging

logger = logging.getLogger(__name__)


KNOWN_FORMATS = {
    "compound_discoverer": ["Name", "Formula", "m/z", "RT", "Area"],
    "lipidsearch": ["LipidMolec", "ClassKey", "FAKey", "CalcMass", "BaseRt", "TotalGrade", "Area"],
    "lipidsearch_alignment": [],
    "el_maven": ["label", "metaGroupId", "groupId", "goodPeakCount", "medMz", "medRt", "maxQuality", "adductName", "isotopeLabel", "compound", "compoundId", "formula", "expectedRtDiff", "ppmDiff", "parent"],
    "compound_discoverer_metadata": ["Sample Identifier", "File", "Sample Type", "Condition"],
    "sample_metadata": ["Sample", "Group"],
}

EL_MAVEN_FEATURE_COLUMNS = {
    "label", "metagroupid", "groupid", "goodpeakcount", "medmz", "medrt", "maxquality",
    "adductname", "isotopelabel", "compound", "compoundid", "formula", "expectedrtdiff",
    "ppmdiff", "parent",
}



def _is_el_maven_df(df: pd.DataFrame) -> bool:
    """Return True if the DataFrame looks like an El-MAVEN isotopologue export."""
    cols = {str(c).lower() for c in df.columns}
    return {"isotopelabel", "compound", "parent"}.issubset(cols) or (
        "isotopelabel" in cols and "compoundid" in cols
    )


def _el_maven_sample_group(col: str) -> str:
    """Derive a group label from an El-MAVEN sample intensity column name."""
    name = str(col).strip()
    # Strip common polarity suffix (e.g. *_pos, *_neg, *pos, *neg).
    name = re.sub(r"(?:^|_)(?:pos|neg)$", "", name, flags=re.I)
    # Explicit replicate marker: GROUP_R1, GROUP_R2, ...
    m = re.match(r"^(?P<group>.+?)_R\d+$", name, re.I)
    if m:
        return m.group("group").strip("_")
    # Trailing replicate index after a letter, e.g. Blank1, Pool1, 250K1.
    m = re.match(r"^(?P<group>.*?[A-Za-z])\d+$", name)
    if m:
        return m.group("group").strip("_")
    return name


def _is_lipidsearch_alignment_file(path: str) -> bool:
    """A LipidSearch AlignmentSettings .txt has section headers like *Parameters setting and *Target search job."""
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as fh:
            for i, line in enumerate(fh):
                if i > 50:
                    break
                if line.startswith("*Parameters setting") or line.startswith("*Target search job"):
                    return True
    except Exception as _exc:
        logger.exception("Unexpected error")
        pass
    return False

SAMPLE_REGEX = re.compile(r"area|intensity|abundance|sample|ctrl|control|treat|rep|replicate|isotopolog|m\+\d+|_[0-9]+$", re.I)
GROUP_REGEX = re.compile(r"^(?P<group>.+?)_(?P<idx>[0-9]+)$")

# Map common Compound Discoverer "Sample Type" values to a clean control/QC group label.
CONTROL_TYPE_LABELS = {
    "blank": "Blank",
    "qc": "QC",
    "quality control": "QC",
    "solvent": "Solvent",
    "standard": "Standard",
    "std": "Standard",
    "pool": "Pool",
    "ntc": "NTC",
}


def _control_group_label(sample_type: str) -> str | None:
    """Return a canonical control/QC group label if sample_type is a QC/Blank/etc."""
    st = str(sample_type).strip().lower()
    return CONTROL_TYPE_LABELS.get(st)


def _sample_group_from_meta(meta: Dict[str, Any], fallback: str) -> str:
    """Choose the best group label from parsed CD metadata.

    Prefer a control/QC label from Sample Type so Blanks and QCs are not
    accidentally merged into a biological Condition (e.g. Non-Exercise).
    """
    sample_type = str(meta.get("sample_type") or "").strip()
    control_label = _control_group_label(sample_type)
    if control_label:
        return control_label
    condition = str(meta.get("condition") or "").strip()
    if condition and condition.lower() not in CONTROL_TYPE_LABELS:
        return condition
    return sample_type or fallback

# LipidSearch 5.x bracketed area columns, e.g. OrgMeanArea[s01], NormArea[s01-1]
LIPIDSEARCH_AREA_RE = re.compile(
    r"^(?P<type>OrgMeanArea|NormArea|OriginalArea|NormHeight|OriginalHeight|Conc)"
    r"\[(?P<sample>s?\d+(?:-\d+)?)\]$",
    re.I,
)

# Compound Discoverer area columns, e.g. "Area: EL-008.raw (F26)"
COMPOUND_DISCOVERER_AREA_RE = re.compile(
    r"^Area:\s*(?P<raw>.+?)\s*\((?P<filecode>F\d+)\)\s*$",
    re.I,
)


def _normalize_lipidsearch_sample_id(sample_id: str) -> str:
    """Convert 's01-1' -> 's1-1' to match alignment file sample ids."""
    if not sample_id:
        return sample_id
    return re.sub(r"(?<![0-9])0+(?=[0-9])", "", sample_id.lower())


def _base_raw_name(raw: str) -> str:
    """Strip common raw-file extensions from a Compound Discoverer raw name."""
    name = str(raw).strip()
    for ext in [".raw", "_raw"]:
        if name.lower().endswith(ext):
            name = name[: -len(ext)]
    return name


def _is_main_row(row: Tuple) -> bool:
    """A Compound Discoverer main compound row starts with True/False in the first cell and has a name in the second."""
    if len(row) < 2:
        return False
    first = row[0]
    if first is None or (isinstance(first, str) and first.strip() == ""):
        return False
    if isinstance(first, str):
        first_clean = first.strip().lower()
        if first_clean in ("true", "1", "1.0"):
            first = True
        elif first_clean in ("false", "0", "0.0"):
            first = False
        else:
            return False
    if isinstance(first, (int, float)) and first in (0, 1):
        pass
    elif not isinstance(first, bool):
        return False
    second = row[1]
    if second is None or (isinstance(second, str) and (second.strip() == "" or second.strip().lower() == "checked")):
        return False
    return True


def _is_lipid_candidate_header(row: Tuple) -> bool:
    """The embedded LipidSearch candidate sub-table header row: empty first cell, 'Checked' in second, 'Name' in fifth."""
    if len(row) < 6:
        return False
    first = row[0]
    if first is not None and not (isinstance(first, str) and first.strip() == ""):
        return False
    second = row[1]
    if not (isinstance(second, str) and second.strip().lower() == "checked"):
        return False
    fifth = row[4]
    if not (isinstance(fifth, str) and fifth.strip().lower() == "name"):
        return False
    return True


def is_compound_discoverer_lipidset(path: str, sheet: str = None) -> bool:
    """Detect a Compound Discoverer export that contains an embedded LipidSearch candidate sub-table."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as _exc:
        logger.exception("Unexpected error")
        return False
    ws = wb[sheet] if sheet else wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if i > 50:
            break
        if _is_lipid_candidate_header(row):
            return True
    return False


def read_compound_discoverer_lipidset(path: str, sheet: str = None) -> pd.DataFrame:
    """Parse a Compound Discoverer + LipidSearch combined export.

    The main compound rows become the data rows; the embedded LipidSearch
    candidate details are attached as a `_lipid_candidates` column per row.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    main_col_idx = {i: str(h).strip() for i, h in enumerate(header) if h is not None}

    main_records: List[Dict[str, Any]] = []
    candidates_by_idx: Dict[int, List[Dict[str, Any]]] = {}
    current_idx: int | None = None
    cand_col_idx: Dict[int, str] = {}

    for row in rows:
        if _is_main_row(row):
            record = {h: (row[i] if i < len(row) else None) for i, h in main_col_idx.items()}
            main_records.append(record)
            current_idx = len(main_records) - 1
            cand_col_idx = {}
            continue

        if _is_lipid_candidate_header(row):
            cand_col_idx = {i: str(h).strip() for i, h in enumerate(row) if h is not None and str(h).strip() != ""}
            continue

        if current_idx is not None and cand_col_idx:
            candidate: Dict[str, Any] = {}
            for i, h in cand_col_idx.items():
                val = row[i] if i < len(row) else None
                if h.lower() == "structure" and isinstance(val, str) and val.startswith("#mol format"):
                    val = None
                candidate[h] = val
            if candidate.get("Name"):
                candidates_by_idx.setdefault(current_idx, []).append(candidate)

    df = pd.DataFrame(main_records)
    if not df.empty:
        df["_lipid_candidates"] = df.index.map(lambda i: candidates_by_idx.get(i, []))
    return df


def select_top_lipid_candidate(candidates: List[Dict[str, Any]]) -> Dict[str, Any] | None:
    """Pick the best LipidSearch candidate using rank, grade, then ID score."""
    if not candidates:
        return None

    def _grade_score(grade):
        if not grade:
            return 99
        g = str(grade).strip().upper()
        order = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4, "F": 5}
        return order.get(g[0], 99)

    scored = []
    for c in candidates:
        rank = c.get("Rank")
        try:
            rank = int(float(rank))
        except Exception as _exc:
            logger.exception("Unexpected error")
            rank = 999
        grade = c.get("Grade") or c.get("Best Grade")
        gs = _grade_score(grade)
        id_score = c.get("ID Score") or c.get("Best ID Score")
        try:
            id_score = float(id_score)
        except Exception as _exc:
            logger.exception("Unexpected error")
            id_score = 0.0
        scored.append(((rank, gs, -id_score), c))

    scored.sort(key=lambda x: x[0])
    return scored[0][1]


def _detect_lipidsearch_samples(columns: List[str]) -> Tuple[List[str], Dict[str, str], Dict[str, List[str]]]:
    """Return preferred sample columns, sample groups, and all candidate intensity sets for LipidSearch exports."""
    candidates: Dict[str, List[str]] = {}
    for col in columns:
        m = LIPIDSEARCH_AREA_RE.match(str(col))
        if m:
            intensity_type = m.group("type").lower()
            candidates.setdefault(intensity_type, []).append(col)

    # Preferred order: normalized area per replicate, then original area, then mean area, then conc
    priority = ["normarea", "originalarea", "orgmeanarea", "conc"]
    chosen = []
    for itype in priority:
        if itype in candidates:
            chosen = candidates[itype]
            break

    sample_groups = {}
    for col in chosen:
        m = LIPIDSEARCH_AREA_RE.match(str(col))
        if m:
            sid = _normalize_lipidsearch_sample_id(m.group("sample"))
            group = sid.split("-")[0]
            sample_groups[col] = group

    return chosen, sample_groups, candidates


def _detect_compound_discoverer_samples(columns: List[str], metadata: Dict[str, Dict[str, Any]]) -> Tuple[List[str], Dict[str, str]]:
    """Detect Compound Discoverer sample area columns and apply metadata mapping."""
    sample_columns = []
    sample_groups = {}
    for col in columns:
        m = COMPOUND_DISCOVERER_AREA_RE.match(str(col))
        if m:
            sample_columns.append(col)
            filecode = m.group("filecode").upper()
            raw_base = _base_raw_name(m.group("raw"))
            meta = None
            if metadata:
                if filecode in metadata:
                    meta = metadata[filecode]
                elif raw_base.lower() in metadata:
                    meta = metadata[raw_base.lower()]
            if meta:
                sample_groups[col] = _sample_group_from_meta(meta, raw_base)
            else:
                sample_groups[col] = raw_base
    return sample_columns, sample_groups


def read_file_to_df(path: str, sheet: str = None) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        if is_compound_discoverer_lipidset(path, sheet):
            return read_compound_discoverer_lipidset(path, sheet)
        if sheet:
            return pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        return pd.read_excel(path, engine="openpyxl")
    if ext == ".csv":
        return pd.read_csv(path, low_memory=False)
    if ext in [".tsv", ".txt"]:
        return pd.read_csv(path, sep="\t", low_memory=False, encoding="utf-8")
    raise ValueError(f"Unsupported file extension: {ext}")


def list_sheets(path: str) -> List[str]:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        xl = pd.ExcelFile(path, engine="openpyxl")
        return xl.sheet_names
    return []


def _detect_metadata_format(columns: List[str]) -> str | None:
    """Classify an uploaded sample metadata / alignment file based on column names."""
    cols_lower = [c.lower() for c in columns]
    col_set = set(cols_lower)
    # Compound Discoverer sample metadata exports
    if {"sample identifier", "file"} <= col_set or {"sample", "file", "sample type"} <= col_set:
        return "compound_discoverer_metadata"
    # Generic sample group metadata
    if "sample" in col_set and any(g in col_set for g in ["group", "condition", "sample type", "sample_type"]):
        return "sample_metadata"
    return None


def detect_file_format(path: str, ext: str = None) -> Dict[str, Any]:
    ext = ext or os.path.splitext(path)[1].lower()
    # LipidSearch AlignmentSettings has section markers before a tabular header.
    if ext in [".txt", ".tsv"] and _is_lipidsearch_alignment_file(path):
        return {"format": "lipidsearch_alignment", "scores": {}, "sheets": [], "columns": []}

    try:
        sheets = list_sheets(path)
        if sheets:
            df = pd.read_excel(path, sheet_name=sheets[0], engine="openpyxl", nrows=5)
        else:
            if ext == ".csv":
                df = pd.read_csv(path, nrows=5, low_memory=False)
            else:
                df = pd.read_csv(path, sep="\t", nrows=5, low_memory=False)
    except Exception as e:
        return {"format": None, "error": str(e), "sheets": []}

    columns = [str(c) for c in df.columns]
    scores = {}
    for fmt, markers in KNOWN_FORMATS.items():
        scores[fmt] = sum(1 for m in markers if any(m.lower() in c.lower() for c in columns))
    best = max(scores, key=scores.get) if scores and max(scores.values()) > 0 else None

    # Fallback: if no data-format markers matched, see if this is a metadata file.
    if not best:
        meta_format = _detect_metadata_format(columns)
        if meta_format:
            best = meta_format

    return {"format": best, "scores": scores, "sheets": sheets, "columns": columns}


def detect_columns(df: pd.DataFrame, metadata: Dict[str, Dict[str, Any]] = None) -> Dict[str, Any]:
    columns = [str(c) for c in df.columns if not str(c).startswith("_")]

    feature_markers = {
        "feature_id": ["name", "compound", "lipidion", "lipid ion", "metabolite", "lipidmolec"],
        "formula": ["formula", "molecular formula"],
        "mz": ["m/z", "precursor mz", "precursor_mz", "calcmass"],
        "rt": ["retention time", "retentiontime", "rt (min)", "rt [min]", "rt", "basert"],
        "adduct": ["adduct", "reference ion", "ion type", "ion_type"],
        "lipid_class": ["lipid class", "lipid_class", "classkey"],
        "grade": ["grade", "score", "totalgrade"],
        "fa": ["fa", "fatty acyl", "fattyacid", "fa composition", "fakey"],
        "calc_mw": ["calc. mw", "calc.mw", "calculated mw"],
        "polarity": ["polarity"],
        "neutral_losses": ["neutral losses", "neutral_losses"],
        "peak_rating": ["peak rating", "peak_rating"],
        "num_lipidsearch_results": ["# lipidsearch results", "lipidsearch results"],
    }
    suggested = {}
    already = set()
    for key, markers in feature_markers.items():
        for col in columns:
            col_lower = col.lower()
            if key == "lipid_class":
                # avoid CD "Class Coverage", "Class FISh Score" columns
                if any(w in col_lower for w in ["coverage", "fish", "score"]):
                    continue
            if any(
                col_lower == m
                or col_lower.startswith(m + " ")
                or col_lower.endswith(" " + m)
                or (m in col_lower and "(" + m + ")" in col_lower)
                for m in markers
            ):
                if col not in already:
                    suggested[key] = col
                    already.add(col)
                    break

    # El-MAVEN isotopologue exports: each row is a compound-isotopologue and the
    # remaining columns are per-sample intensities (e.g. *_R1_pos).
    if _is_el_maven_df(df):
        feature_cols = {c for c in columns if str(c).lower() in EL_MAVEN_FEATURE_COLUMNS}
        already.update(feature_cols)
        sample_columns = [c for c in columns if c not in already]
        sample_groups = {c: _el_maven_sample_group(c) for c in sample_columns}
        suggested.setdefault("feature_id", "compound")
        suggested.setdefault("formula", "formula")
        if "medMz" in columns:
            suggested.setdefault("mz", "medMz")
        if "medRt" in columns:
            suggested.setdefault("rt", "medRt")
        if "adductName" in columns:
            suggested.setdefault("adduct", "adductName")
        return {
            "suggested_mapping": suggested,
            "sample_columns": sample_columns,
            "feature_columns": [c for c in columns if c in feature_cols],
            "sample_groups": sample_groups,
            "lipidsearch_candidates": {},
        }

    # Try Compound Discoverer-specific area columns first
    cd_samples, cd_groups = _detect_compound_discoverer_samples(columns, metadata or {})

    # Try LipidSearch-specific area columns next
    ls_samples, ls_groups, ls_candidates = _detect_lipidsearch_samples(columns)

    if cd_samples:
        sample_columns = cd_samples
        sample_groups = cd_groups
    elif ls_samples:
        sample_columns = ls_samples
        sample_groups = ls_groups
    else:
        sample_columns = [c for c in columns if SAMPLE_REGEX.search(str(c)) and c not in already]
        if not sample_columns:
            sample_columns = [c for c in columns if re.search(r"_[0-9]+$", str(c))]

        sample_groups = {}
        for col in sample_columns:
            match = GROUP_REGEX.match(str(col))
            if match:
                sample_groups[col] = match.group("group")
            else:
                sample_groups[col] = "unknown"

    # Apply LipidSearch alignment file mapping if metadata was supplied that way
    if metadata:
        for col in sample_columns:
            m = LIPIDSEARCH_AREA_RE.match(str(col))
            if m:
                sid = _normalize_lipidsearch_sample_id(m.group("sample"))
                if sid in metadata:
                    sample_groups[col] = metadata[sid].get("condition", sample_groups[col])
            elif col in metadata:
                sample_groups[col] = metadata[col].get("condition", sample_groups[col])

    feature_columns = [c for c in columns if c not in sample_columns and c not in already]
    return {
        "suggested_mapping": suggested,
        "sample_columns": sample_columns,
        "feature_columns": feature_columns,
        "sample_groups": sample_groups,
        "lipidsearch_candidates": ls_candidates,
    }


def parse_lipidsearch_alignment(path: str) -> Dict[str, Dict[str, Any]]:
    """Parse a LipidSearch alignment .txt file and return sample_id/raw -> metadata mapping."""
    mapping: Dict[str, Dict[str, Any]] = {}
    if not os.path.exists(path):
        return mapping

    with open(path, "r", encoding="utf-8", errors="ignore") as fh:
        in_section = False
        for raw_line in fh:
            line = raw_line.replace("\r\n", "\n").replace("\r", "\n").rstrip("\n")
            if not line:
                continue
            if line.startswith("*Target search job"):
                in_section = True
                continue
            if in_section:
                if line.startswith("*"):
                    break
                parts = line.split("\t")
                if len(parts) >= 4:
                    sid = _normalize_lipidsearch_sample_id(parts[2].strip())
                    raw = _base_raw_name(parts[1].strip())
                    # LipidSearch marks the first replicate of each group with a leading '*'.
                    condition = parts[3].strip().lstrip("*").strip()
                    # Canonicalize QC/Blank/Pool labels for downstream QC logic.
                    control_label = _control_group_label(condition)
                    if control_label:
                        condition = control_label
                    meta = {"condition": condition, "raw_name": raw, "sample_identifier": sid}
                    mapping[sid] = meta
                    if raw:
                        mapping[raw.lower()] = meta
                        mapping[raw] = meta
    return mapping


def parse_compound_discoverer_metadata(path: str) -> Dict[str, Dict[str, Any]]:
    """Parse a Compound Discoverer metadata .xlsx/.csv/.txt and return file identifier -> metadata mapping."""
    mapping: Dict[str, Dict[str, Any]] = {}
    if not os.path.exists(path):
        return mapping

    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in [".xlsx", ".xls"]:
            xl = pd.ExcelFile(path, engine="openpyxl")
            df = pd.read_excel(path, sheet_name=xl.sheet_names[0])
        elif ext == ".csv":
            df = pd.read_csv(path)
        else:
            df = pd.read_csv(path, sep="\t")
    except Exception as _exc:
        logger.exception("Unexpected error")
        return mapping

    for _, row in df.iterrows():
        def _get(col):
            val = row.get(col, "")
            if pd.isna(val):
                return ""
            return str(val).strip()

        sample = _get("Sample")
        file_code = _get("File").upper()
        sample_id = _get("Sample Identifier")
        sample_type = _get("Sample Type")
        condition = _get("Condition")

        meta = {
            "sample": sample,
            "file_code": file_code,
            "sample_identifier": sample_id,
            "sample_type": sample_type,
            "condition": condition or sample_type or sample_id,
        }
        if file_code:
            mapping[file_code] = meta
        if sample_id:
            mapping[_base_raw_name(sample_id).lower()] = meta
            mapping[sample_id.lower()] = meta
        if sample:
            mapping[sample.upper()] = meta
    return mapping


def parse_sample_metadata(path: str) -> Dict[str, Dict[str, Any]]:
    """Dispatch to the correct metadata/alignment parser based on file content."""
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls", ".csv"]:
        # Try CD metadata first; if empty, fall back to generic
        mapping = parse_compound_discoverer_metadata(path)
        if mapping:
            return mapping
    if ext in [".txt", ".tsv", ".csv"]:
        mapping = parse_lipidsearch_alignment(path)
        if mapping:
            return mapping
    return {}


async def preview_file(uploaded, sheet: str = None, metadata: Dict[str, Dict[str, Any]] = None, db=None) -> Dict[str, Any]:
    path = await storage.get_file_path(uploaded.stored_name, db) if db else os.path.join(settings.UPLOAD_DIR, uploaded.stored_name)
    df = read_file_to_df(path, sheet)
    mapping = detect_columns(df, metadata=metadata)
    return {
        "detected_format": uploaded.detected_format,
        "sheets": list_sheets(path),
        "columns": [str(c) for c in df.columns],
        "sample_columns": mapping["sample_columns"],
        "feature_columns": mapping["feature_columns"],
        "row_count": len(df),
        "suggested_mapping": mapping["suggested_mapping"],
        "sample_groups": mapping["sample_groups"],
        "lipidsearch_candidates": mapping.get("lipidsearch_candidates", {}),
    }
